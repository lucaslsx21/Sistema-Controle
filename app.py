import customtkinter as ctk
from customtkinter import *
import openpyxl
import pathlib
from tkinter import messagebox
from openpyxl import Workbook
from datetime import date as dt


class BackEnd():
    def todo_sistema(self):
        self.ficheiro = pathlib.Path(r"ControloDocumentos.xlsx")
        if self.ficheiro.exists():
            pass
        else:
            pass

            # folha de remuneracao
            self.ficheiro = Workbook()
            self.folha = self.ficheiro.active
            self.folha.title = "Remuneracao"
            self.folha["A1"] = "Nome"
            self.folha["B1"] = "Data"
            self.folha["C1"] = "Sigla"
            self.folha["D1"] = "Curso"
            self.folha["E1"] = "Unidade Curricular"
            self.folha["F1"] = "Ano"
            self.folha["G1"] = "Semestre"
            self.folha["H1"] = "Carga Horaria"
            self.folha["I1"] = "Obs. Adicionais"

            # folha de atividades
            self.folha1 = self.ficheiro.create_sheet("Atividades")
            self.folha1["A1"] = "Data"
            self.folha1["B1"] = "Nome"
            self.folha1["C1"] = "Sigla"
            self.folha1["D1"] = "Area"
            self.folha1["E1"] = "Seccao"
            self.folha1["F1"] = "Curso"
            self.folha1["G1"] = "Ano"
            self.folha1["H1"] = "Atividade"
            self.folha1["I1"] = "Carga Horaria"
            self.folha1["J1"] = "Obs. Adicionais"

            # folha geral
            self.folha2 = self.ficheiro.create_sheet("Geral")
            self.folha2["A1"] = "Data"
            self.folha2["B1"] = "Nome"
            self.folha2["C1"] = "Sigla"
            self.folha2["D1"] = "Area"
            self.folha2["E1"] = "Seccao"
            self.folha2["F1"] = "Assunto"
            self.folha2["G1"] = "Obs. Adicionais"

            self.ficheiro.save(r"ControloDocumentos.xlsx")

# =========================SALVAR REMUNERACAO===============================
    def salvar_remuneracao(self):
        # Pegar os dados que estao a entrar atraves do formulario (dados do usuario)
        self.hoje = dt.today()
        self.data = self.hoje.strftime("%d/%m/%Y")
        self.nome = self.r_nome_entry.get()
        self.sigla = self.r_sigla_entry.get()
        self.uc = self.r_uc_entry.get()
        self.horas_lecionadas = int(self.r_horas_lecionadas_entry.get())
        self.curso = self.r_curso_opt.get()
        self.ano = self.r_ano_opt.get()
        self.semestre = self.r_semestre_opt.get()
        self.carga_horaria = int(self.r_carga_horaria_opt.get())
        self.obs = self.r_observacoes.get("0.0", "end")

        # Abrir o ficheiro e a folha e localizar a folha
        self.ficheiro = openpyxl.load_workbook(r"ControloDocumentos.xlsx")
        self.folha = self.ficheiro.get_sheet_by_name(r"Remuneracao")

        # Colocar cada informacao na sua célula
        self.folha.cell(column=1, row=self.folha.max_row+1, value=self.data)
        self.folha.cell(column=2, row=self.folha.max_row, value=self.nome)
        self.folha.cell(column=3, row=self.folha.max_row, value=self.sigla)
        self.folha.cell(column=4, row=self.folha.max_row, value=self.curso)
        self.folha.cell(column=5, row=self.folha.max_row, value=self.uc)
        self.folha.cell(column=6, row=self.folha.max_row, value=self.ano)
        self.folha.cell(column=7, row=self.folha.max_row, value=self.semestre)
        self.folha.cell(column=8, row=self.folha.max_row,
                        value=self.carga_horaria)
        self.folha.cell(column=9, row=self.folha.max_row,
                        value=self.horas_lecionadas)
        self.folha.cell(column=10, row=self.folha.max_row, value=self.obs)

        # Salvar os dados dentro do sistema
        self.ficheiro.save(r"ControloDocumentos.xlsx")
        messagebox.showinfo("Sistema", "Dados guardados com sucesso :) !")

# =============================SALVAR ATIVIDADES==========================

    def salvar_atividades(self):
        # Pegar os dados que estao a entrar atraves do formulario (dados do usuario)
        self.hoje = dt.today()
        self.data = self.hoje.strftime("%d/%m/%Y")
        self.nome = self.a_nome_entry.get()
        self.sigla = self.a_sigla_entry.get()
        self.area = self.a_area_entry.get()
        self.secao = self.a_seccao_opt.get()
        self.curso = self.a_curso_opt.get()
        self.ano = self.a_ano_opt.get()
        self.atividade = self.a_atividade_opt.get()
        self.carga_horaria = int(self.a_carga_horaria_opt.get())
        self.obs = self.a_observacoes.get("0.0", "end")

        # Abrir o ficheiro e a folha e localizar a folha
        self.ficheiro = openpyxl.load_workbook(r"ControloDocumentos.xlsx")
        self.folha1 = self.ficheiro.get_sheet_by_name(r"Atividades")

        # Colocar cada informacao na sua célula
        self.folha1.cell(column=1, row=self.folha1.max_row+1, value=self.data)
        self.folha1.cell(column=2, row=self.folha1.max_row, value=self.nome)
        self.folha1.cell(column=3, row=self.folha1.max_row, value=self.sigla)
        self.folha1.cell(column=4, row=self.folha1.max_row, value=self.area)
        self.folha1.cell(column=5, row=self.folha1.max_row, value=self.secao)
        self.folha1.cell(column=6, row=self.folha1.max_row, value=self.curso)
        self.folha1.cell(column=7, row=self.folha1.max_row, value=self.ano)
        self.folha1.cell(column=8, row=self.folha1.max_row,
                         value=self.atividade)
        self.folha1.cell(column=9, row=self.folha1.max_row,
                         value=self.carga_horaria)
        self.folha1.cell(column=10, row=self.folha1.max_row, value=self.obs)

        # Salvar os dados dentro do sistema
        self.ficheiro.save(r"ControloDocumentos.xlsx")
        messagebox.showinfo("Sistema", "Dados guardados com sucesso :) !")


# =======================SALVAR GERAL=======================================

    def salvar_geral(self):
        # Pegar os dados que estao a entrar atraves do formulario (dados do usuario)
        self.hoje = dt.today()
        self.data = self.hoje.strftime("%d/%m/%Y")
        self.nome = self.g_nome_entry.get()
        self.sigla = self.g_sigla_entry.get()
        self.area = self.g_area_entry.get()
        self.secao = self.g_seccao_opt.get()
        self.assunto = self.g_assunto_entry.get()
        self.obs = self.g_observacoes.get("0.0", "end")

        # Abrir o ficheiro e a folha e localizar a folha
        self.ficheiro = openpyxl.load_workbook(r"ControloDocumentos.xlsx")
        self.folha2 = self.ficheiro.get_sheet_by_name(r"Geral")

        # Colocar cada informacao na sua célula
        self.folha2.cell(column=1, row=self.folha2.max_row+1, value=self.data)
        self.folha2.cell(column=2, row=self.folha2.max_row, value=self.nome)
        self.folha2.cell(column=3, row=self.folha2.max_row, value=self.sigla)
        self.folha2.cell(column=4, row=self.folha2.max_row, value=self.area)
        self.folha2.cell(column=5, row=self.folha2.max_row, value=self.secao)
        self.folha2.cell(column=6, row=self.folha2.max_row, value=self.assunto)
        self.folha2.cell(column=7, row=self.folha2.max_row, value=self.obs)

        # Salvar os dados dentro do sistema
        self.ficheiro.save(r"ControloDocumentos.xlsx")
        messagebox.showinfo("Sistema", "Dados guardados com sucesso :) !")


class App(ctk.CTk, BackEnd):
    def __init__(self):
        super().__init__()
        self.windows_config()
        self.tema_config()
        self.todo_sistema()
        self.frontend()

    def windows_config(self):
        self.title("SISTEMA DE REMUNERAÇÃO")
        self.geometry("800x650")
        self._set_appearance_mode("System")
        self.minsize(width=800, height=650)

    def tema_config(self):
        self.switch_var = ctk.StringVar(value="off")

        def set_tema():
            if self.switch_var.get() == 'dark':
                ctk.set_appearance_mode("Dark")

            elif self.switch_var.get() == 'light':
                ctk.set_appearance_mode("Light")

            else:
                ctk.set_appearance_mode("System")

        self.switch = ctk.CTkSwitch(
            self, text="TEMA", variable=self.switch_var, onvalue="dark", offvalue="light", command=set_tema)
        self.switch.place(x=25, y=20)

    def frontend(self):
        self.tt = ctk.CTkLabel(
            self, text="Sistema de Relatório", font=("arial bold", 30))
        self.tt.pack(pady=20)

        self.tabview = ctk.CTkTabview(self,
                                      width=700,
                                      height=500,
                                      corner_radius=8,
                                      border_width=2,
                                      segmented_button_fg_color="grey",
                                      segmented_button_selected_color="#037",
                                      segmented_button_unselected_hover_color="#111",
                                      segmented_button_unselected_color="#333",)
        self.tabview.pack(pady=10)
        self.tabview.add("Remuneracao")
        self.tabview.add("Atividades")
        self.tabview.add("Geral")

        self.tabview.tab("Remuneracao").grid_columnconfigure(1, weight=1)
        self.tabview.tab("Atividades").grid_columnconfigure(1, weight=1)
        self.tabview.tab("Geral").grid_columnconfigure(1, weight=1)

        # CRIANDOS A TELA DE MATRICULADOS - WINDGETS
        self.trm = ctk.CTkLabel(self.tabview.tab(
            "Remuneracao"), text="REQUERIMENTO DE MATRICULA", font=("arial bold", 20))
        self.nome = ctk.CTkLabel(self.tabview.tab(
            "Remuneracao"), text="Nome:", font=("arial bold", 14))
        self.sigla = ctk.CTkLabel(self.tabview.tab(
            "Remuneracao"), text="Sigla:", font=("arial bold", 14))
        self.uc = ctk.CTkLabel(self.tabview.tab(
            "Remuneracao"), text="Unidade Curricular:", font=("arial bold", 14))
        self.horas_lecionadas = ctk.CTkLabel(self.tabview.tab(
            "Remuneracao"), text="Horas Lecionadas", font=("arial bold", 14))
        self.obs = ctk.CTkLabel(self.tabview.tab(
            "Remuneracao"), text="Observacoes:", font=("arial bold", 14))

        # WINDGETS
        self.r_nome_entry = ctk.CTkEntry(self.tabview.tab(
            "Remuneracao"), width=450, height=35, font=("arial", 16))
        self.r_sigla_entry = ctk.CTkEntry(self.tabview.tab(
            "Remuneracao"), width=150, height=35, font=("arial", 16))
        self.r_uc_entry = ctk.CTkEntry(self.tabview.tab(
            "Remuneracao"), width=450, height=35, font=("arial", 16))
        self.r_horas_lecionadas_entry = ctk.CTkEntry(self.tabview.tab(
            "Remuneracao"), width=150, height=35, font=("arial", 16))

        self.r_curso_opt = ctk.CTkComboBox(self.tabview.tab("Remuneracao"),
                                           values=[
            "Soldados de Cristo", "As fiéis", "Crescendo com Cristo"],
            fg_color="#333", width=150, height=35)
        self.r_curso_opt.set("Cursos")

        self.r_ano_opt = ctk.CTkOptionMenu(self.tabview.tab("Remuneracao"),
                                           values=["1 Ano", "2 Ano", "3 Ano"],
                                           fg_color="#333", width=130, height=35)
        self.r_ano_opt.set("Ano")

        self.r_semestre_opt = ctk.CTkOptionMenu(self.tabview.tab("Remuneracao"),
                                                values=[
            "1 Simestre", "2 Simestre", "3 Simestre"],
            fg_color="#333", width=130, height=35)
        self.r_semestre_opt.set("Simestre")

        self.r_carga_horaria_opt = ctk.CTkOptionMenu(self.tabview.tab("Remuneracao"),
                                                     values=["20", "40",
                                                     "50", "60", "100"],
                                                     fg_color="#333", width=150, height=35)
        self.r_carga_horaria_opt.set("Carga Horaria")
        self.r_observacoes = ctk.CTkTextbox(self.tabview.tab(
            "Remuneracao"), width=620, height=70, border_spacing=10, activate_scrollbars=True, font=("arial", 16))
        self.r_save_btn = ctk.CTkButton(self.tabview.tab("Remuneracao"), text="Salvar Dados".upper(
        ), font=("arial bold", 14), fg_color="#037", hover_color="#026", command=self.salvar_remuneracao)

        # POSICIONANDOS OS WINDGETS NA TELA
        self.trm.pack()
        self.nome.place(x=25, y=50)
        self.r_nome_entry.place(x=25, y=80)
        self.sigla.place(x=495, y=50)
        self.r_sigla_entry.place(x=495, y=80)
        self.uc.place(x=25, y=130)
        self.r_uc_entry.place(x=25, y=160)
        self.horas_lecionadas.place(x=495, y=130)
        self.r_horas_lecionadas_entry.place(x=495, y=160)
        self.r_curso_opt.place(x=25, y=210)
        self.r_ano_opt.place(x=195, y=210)
        self.r_semestre_opt.place(x=345, y=210)
        self.r_carga_horaria_opt.place(x=495, y=210)
        self.obs.place(x=25, y=260)
        self.r_observacoes.place(x=25, y=300)
        self.r_save_btn.place(x=280, y=395)

        # ======= FORMULARIO DE ATIVIDADES ======== #
        self.trm = ctk.CTkLabel(self.tabview.tab(
            "Atividades"), text="REQUERIMENTO DE ATIVIDADES", font=("arial bold", 20))
        self.nome = ctk.CTkLabel(self.tabview.tab(
            "Atividades"), text="Nome:", font=("arial bold", 14))
        self.sigla = ctk.CTkLabel(self.tabview.tab(
            "Atividades"), text="Sigla:", font=("arial bold", 14))
        self.area = ctk.CTkLabel(self.tabview.tab(
            "Atividades"), text="Area:", font=("arial bold", 14))
        self.obs = ctk.CTkLabel(self.tabview.tab(
            "Atividades"), text="Observacoes:", font=("arial bold", 14))

        # WIDGETS
        self.a_nome_entry = ctk.CTkEntry(self.tabview.tab(
            "Atividades"), width=450, height=35, font=("arial", 16))
        self.a_sigla_entry = ctk.CTkEntry(self.tabview.tab(
            "Atividades"), width=150, height=35, font=("arial", 16))
        self.a_area_entry = ctk.CTkEntry(self.tabview.tab(
            "Atividades"), width=450, height=35, font=("arial", 16))
        self.a_horas_lecionadas_entry = ctk.CTkEntry(self.tabview.tab(
            "Atividades"), width=150, height=35, font=("arial", 16))

        self.a_curso_opt = ctk.CTkComboBox(self.tabview.tab("Atividades"),
                                           values=[
            "Soldados de Cristo", "As fiéis", "Crescendo com Cristo"],
            fg_color="#333", width=150, height=35)
        self.a_curso_opt.set("Cursos")

        self.a_ano_opt = ctk.CTkOptionMenu(self.tabview.tab("Atividades"),
                                           values=["1 Ano", "2 Ano", "3 Ano"],
                                           fg_color="#333", width=130, height=35)
        self.a_ano_opt.set("Ano")

        self.a_atividade_opt = ctk.CTkOptionMenu(self.tabview.tab("Atividades"),
                                                 values=[
                                                     "Correcao de Avaliacao", "Juri de Memorias", "Orientacao de estagio", "Outras atividades"],
                                                 fg_color="#333", width=130, height=35)
        self.a_atividade_opt.set("Atividade")

        self.a_seccao_opt = ctk.CTkOptionMenu(self.tabview.tab("Atividades"),
                                              values=[
            "Sold", "AsF", "Cres"],
            fg_color="#333", width=150, height=35)
        self.a_seccao_opt.set("Seccao")

        self.a_carga_horaria_opt = ctk.CTkOptionMenu(self.tabview.tab("Atividades"),
                                                     values=["1", "2",
                                                     "3", "5", "10"],
                                                     fg_color="#333", width=150, height=35)
        self.a_carga_horaria_opt.set("Carga Horaria")
        self.a_observacoes = ctk.CTkTextbox(self.tabview.tab(
            "Atividades"), width=620, height=70, border_spacing=10, activate_scrollbars=True, font=("arial", 16))
        self.a_save_btn = ctk.CTkButton(self.tabview.tab("Atividades"), text="Salvar Dados".upper(
        ), font=("arial bold", 14), fg_color="#037", hover_color="#026", command=self.salvar_atividades)

        # POSICIONANDO OS WINDGETS NA TELA
        self.trm.pack()
        self.nome.place(x=25, y=50)
        self.a_nome_entry.place(x=25, y=80)
        self.sigla.place(x=495, y=50)
        self.a_sigla_entry.place(x=495, y=80)
        self.area.place(x=25, y=130)
        self.a_area_entry.place(x=25, y=160)
        self.a_seccao_opt.place(x=495, y=160)
        self.a_curso_opt.place(x=25, y=210)
        self.a_ano_opt.place(x=195, y=210)
        self.a_atividade_opt.place(x=345, y=210)
        self.a_carga_horaria_opt.place(x=495, y=210)
        self.obs.place(x=25, y=260)
        self.a_observacoes.place(x=25, y=300)
        self.a_save_btn.place(x=280, y=395)

        # ======= FORMULARIO DE DOCUMENTOS GERAIS ======== #
        self.trm = ctk.CTkLabel(self.tabview.tab(
            "Geral"), text="REQUERIMENTO DE GERAL".upper(), font=("arial bold", 20))
        self.nome = ctk.CTkLabel(self.tabview.tab(
            "Geral"), text="Nome:", font=("arial bold", 14))
        self.sigla = ctk.CTkLabel(self.tabview.tab(
            "Geral"), text="Sigla:", font=("arial bold", 14))
        self.area = ctk.CTkLabel(self.tabview.tab(
            "Geral"), text="Area:", font=("arial bold", 14))
        self.obs = ctk.CTkLabel(self.tabview.tab(
            "Geral"), text="Observacoes:", font=("arial bold", 14))
        self.assunto = ctk.CTkLabel(self.tabview.tab(
            "Geral"), text="Assunto:", font=("arial bold", 14))

        # WIDGETS
        self.g_nome_entry = ctk.CTkEntry(self.tabview.tab(
            "Geral"), width=450, height=35, font=("arial", 16))
        self.g_sigla_entry = ctk.CTkEntry(self.tabview.tab(
            "Geral"), width=150, height=35, font=("arial", 16))
        self.g_area_entry = ctk.CTkEntry(self.tabview.tab(
            "Geral"), width=450, height=35, font=("arial", 16))
        self.g_horas_lecionadas_entry = ctk.CTkEntry(self.tabview.tab(
            "Geral"), width=150, height=35, font=("arial", 16))
        self.g_assunto_entry = ctk.CTkEntry(self.tabview.tab(
            "Geral"), width=620, height=35, font=("arial", 16))

        self.g_seccao_opt = ctk.CTkOptionMenu(self.tabview.tab("Geral"),
                                              values=[
            "Sold", "AsF", "Cres"],
            fg_color="#333", width=150, height=35)
        self.g_seccao_opt.set("Seccao")

        self.g_observacoes = ctk.CTkTextbox(self.tabview.tab(
            "Geral"), width=620, height=70, border_spacing=10, activate_scrollbars=True, font=("arial", 16))
        self.g_save_btn = ctk.CTkButton(self.tabview.tab("Geral"), text="Salvar Dados".upper(
        ), font=("arial bold", 14), fg_color="#037", hover_color="#026", command=self.salvar_geral)

        # POSICIONANDO OS WINDGETS NA TELA
        self.trm.pack()
        self.nome.place(x=25, y=50)
        self.g_nome_entry.place(x=25, y=80)
        self.sigla.place(x=495, y=50)
        self.g_sigla_entry.place(x=495, y=80)
        self.area.place(x=25, y=120)
        self.g_area_entry.place(x=25, y=150)
        self.g_seccao_opt.place(x=495, y=150)
        self.assunto.place(x=25, y=190)
        self.g_assunto_entry.place(x=25, y=220)
        self.obs.place(x=25, y=260)
        self.g_observacoes.place(x=25, y=300)
        self.g_save_btn.place(x=280, y=395)


if __name__ == "__main__":
    app = App()
    app.mainloop()
